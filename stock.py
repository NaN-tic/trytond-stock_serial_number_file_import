# The COPYRIGHT file at the top level of this repository contains the full
# copyright notices and license terms.
from openpyxl import load_workbook
from zipfile import ZipFile
try:
    import cStringIO as StringIO
except ImportError:
    import StringIO


from trytond.model import fields
from trytond.pool import Pool, PoolMeta
from trytond.transaction import Transaction

__all__ = ['SplitMoveStart', 'SplitMove']
__metaclass__ = PoolMeta


class SplitMoveStart:
    __name__ = 'stock.move.split.start'

    lot_file = fields.Binary('File')


class SplitMove:
    __name__ = 'stock.move.split'

    def transition_split(self):
        pool = Pool()
        Move = pool.get('stock.move')
        Lot = pool.get('stock.lot')
        move = Move(Transaction().context['active_id'])
        lots = self.start.lots
        if self.start.lot_file:
            to_create = []
            book = load_workbook(StringIO.StringIO(self.start.lot_file),
                use_iterators=True)
            sheet = book.worksheets[0]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.internal_value:
                        current_lots = Lot.search([
                                ('product', '=', move.product),
                                ('number', '=', str(cell.internal_value)),
                                ], limit=1)
                        if current_lots:
                            lots.append(current_lots[0])
                            continue
                        to_create.append({
                                'number': str(cell.internal_value),
                                'product': move.product.id,
                                })
                    break
            if to_create:
                lots = Lot.create(to_create)
        move.split_by_lot(self.start.quantity, self.start.uom,
            count=self.start.count, lots=lots, start_lot=self.start.start_lot,
            end_lot=self.start.end_lot)
        return 'end'
